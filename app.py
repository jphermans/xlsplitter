import streamlit as st
import tempfile
import os
import zipfile
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation

st.set_page_config(page_title="Excel Splitter", layout="centered")
st.title("📊 Excel Splitter with Validation & Formatting")

uploaded_file = st.file_uploader("📥 Upload an Excel file (.xlsx)", type=["xlsx"])
rows_per_batch = st.number_input("🔢 Rows per batch", min_value=10, max_value=10000, value=1000, step=100)

if uploaded_file and st.button("🚀 Start splitting"):
    st.caption("Developed by JPHsystems")
    with tempfile.TemporaryDirectory() as temp_dir:
        input_path = os.path.join(temp_dir, "input.xlsx")
        with open(input_path, "wb") as f:
            f.write(uploaded_file.read())

        wb = load_workbook(input_path, data_only=False)
        # Unhide any hidden or very hidden sheets so they are processed like regular sheets
        for sheet in wb.worksheets:
            if sheet.sheet_state != 'visible':
                sheet.sheet_state = 'visible'
        ws = wb.active

        total_rows = ws.max_row
        max_col = ws.max_column
        total_batches = (total_rows - 1) // rows_per_batch + 1

        validations = list(ws.data_validations.dataValidation)
        col_widths = {col: ws.column_dimensions[col].width for col in ws.column_dimensions}

        output_paths = []
        progress_bar = st.progress(0, text="🔄 Batch processing started...")

        for batch_num in range(total_batches):
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = f"Batch_{batch_num + 1}"

            # Copy headers
            for col in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col)
                new_cell = new_ws.cell(row=1, column=col, value=cell.value)
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format

            # Copy column widths
            for col_letter, width in col_widths.items():
                new_ws.column_dimensions[col_letter].width = width

            # Copy rows
            start = batch_num * rows_per_batch + 2
            end = min(start + rows_per_batch - 1, total_rows)

            for row in range(start, end + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    new_cell = new_ws.cell(row=row - start + 2, column=col, value=cell.value)
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.border = copy(cell.border)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.number_format = cell.number_format

            # Add validations
            for dv in validations:
                dv_copy = DataValidation(
                    type=dv.type,
                    formula1=dv.formula1,
                    formula2=dv.formula2,
                    showDropDown=dv.showDropDown,
                    allow_blank=dv.allow_blank,
                    operator=dv.operator,
                    showErrorMessage=dv.showErrorMessage,
                    error=dv.error,
                    errorTitle=dv.errorTitle,
                    prompt=dv.prompt,
                    promptTitle=dv.promptTitle
                )
                for rng in dv.ranges.ranges:
                    dv_copy.add(rng.coord)
                new_ws.add_data_validation(dv_copy)

            # Save batch file
            output_file = os.path.join(temp_dir, f"Batch_{batch_num + 1}.xlsx")
            new_wb.save(output_file)
            output_paths.append(output_file)

            # Update progress
            progress_percent = (batch_num + 1) / total_batches
            progress_bar.progress(progress_percent, text=f"📦 Batch {batch_num + 1}/{total_batches} processed")

        # Create ZIP
        zip_path = os.path.join(temp_dir, "batches.zip")
        with zipfile.ZipFile(zip_path, "w") as zipf:
            for path in output_paths:
                zipf.write(path, arcname=os.path.basename(path))

        # Download button
        with open(zip_path, "rb") as f:
            st.download_button("📦 Download all batches (ZIP)", f.read(), file_name="batches.zip")

        st.success(f"✅ Split into {total_batches} batches!")
        progress_bar.empty()